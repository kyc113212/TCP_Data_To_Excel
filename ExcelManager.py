import os
from PyQt5.QtCore import *
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

class ExcelManager():

    def __init__(self):
        super().__init__()
        self.file_name = ''
        self.base_dir = os.getcwd()
        self.write_wb = Workbook()
        self.write_ws = None
        self.line_cnt = 1

    def file_name_set(self, file_name):
        self.file_name = file_name
        self.write_ws = self.write_wb.create_sheet(file_name)

        for i in range(27):
            self.write_ws.cell(self.line_cnt, i + 4, i + 1)
            self.write_ws.cell(self.line_cnt, i + 4).fill = PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')
        self.line_cnt += 1

    def write_data(self, time_line, rx_data):
        col_idx = 4
        for idx in range(len(rx_data)):
            if idx % 2 == 0:
                continue
            else:
                date_str = self.file_name[0:2] + '년 ' + self.file_name[2:4] + '월 ' + self.file_name[4:6] + '일'
                time_str = ''
                for t in range(len(time_line)):
                    if t == 0:
                        time_str += (time_line[t].lstrip() + ':')
                    elif t == len(time_line) - 1:
                        time_str += time_line[t]
                    else:
                        time_str += (time_line[t] + ':')
                self.write_ws.cell(self.line_cnt, 1, self.line_cnt - 1)
                self.write_ws.cell(self.line_cnt, 1).fill = PatternFill(start_color='99ffb3', end_color='99ffb3', fill_type='solid')
                self.write_ws.cell(self.line_cnt, 2, date_str)
                self.write_ws.cell(self.line_cnt, 3, time_str)
                self.write_ws.cell(self.line_cnt, col_idx, (int(rx_data[idx], 16) / 10))
                col_idx += 1
        self.line_cnt += 1

    def save_excel(self):
        self.write_wb.save("TCP Data.xlsx")

    def init_val(self):
        self.file_name = ''
        self.write_ws = None
        self.line_cnt = 1